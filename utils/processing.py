from langchain.prompts import ChatPromptTemplate
from langchain_openai.chat_models import ChatOpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from io import BytesIO
import time

import streamlit as st

def get_qa_chain(openai_api_key, ai_model_name, model_max_tokens):
    qa_system_prompt = "Analyze the provided data tables to augment the target table with the specified columns. Ensure the output strictly adheres to the requested format, including only the required information in the output."

    qa_input_prompt = """Imagine you are a business Analyst who is an expert in Data Migration Projects. Your client has provided with some claims data that is managed by two TPA's. The first TPA stores the data in this format.

    **Source1 Table:**
    {source1_table}

    **Source2 Table:**
    {source2_table}

    Now you need to map the source data to target and fill the column "Source1 Column Name" and "Source2 Column Name" based on target column "Target Column Name".

    **Target Table:**
    {target_table}"""
    
    qa_prompt = ChatPromptTemplate.from_messages([
        ("system", qa_system_prompt),
        ("human", qa_input_prompt),
    ])

    return qa_prompt | ChatOpenAI(
        model=ai_model_name, 
        streaming=True,
        temperature=0.3, 
        max_tokens=model_max_tokens, 
        openai_api_key=openai_api_key, 
        model_kwargs={
            "top_p": 1,
            "frequency_penalty": 0,
            "presence_penalty": 0,
        }
    )

# utils/processing.py

def process_target_table(
    source1_table_markdown,
    source2_table_markdown, 
    target_table_markdowns, 
    llm,
    target_table_chunk_size
):
    combined_df = pd.DataFrame()
    progress_text = "Please wait..., this may take a moment depending on file size and content."
    progress_bar = st.progress(0, text=progress_text)

    count = 1
    for markdown in target_table_markdowns:
        input_vars = {
            "source1_table": source1_table_markdown,
            "source2_table": source2_table_markdown,
            "target_table": markdown,
        }
        res = llm.invoke(input_vars)
        data = res.content
        lines = data.split('\n')[2:]
        rows = []
        for line in lines:
            columns = [x.strip() for x in line.split('|') if x.strip()]
            row = {
                'Target Column Name': columns[0] if len(columns) >= 1 else "-",
                'Target Column DataType': columns[1] if len(columns) >= 2 else "-",
                'Source1 Column Name': columns[2] if len(columns) >= 3 else "-",
                'Source1 Mapping': columns[3] if len(columns) >= 4 else "-",
                'Source2 Column Name': columns[4] if len(columns) >= 5 else "-",
                'Source2 Mapping': columns[5] if len(columns) == 6 else "-",
            }
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        combined_df = pd.concat([combined_df, df], ignore_index=True)
        count += 1
        progress_num = count / (len(target_table_markdowns))
        progress_bar.progress(1 if progress_num > 1 else progress_num, text=progress_text)

        # Sleep for openai rate limiting
        time.sleep(1)
    
    return combined_df, progress_bar


def get_column_width(text, min_header_weight):
    return max(len(text), min_header_weight) * 1.1

def create_excel_file(
    combined_df, session_state, excel_title_height, excel_header_height, 
    excel_data_height, max_header_weight, min_header_weight, target_filename
):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False

    # Styling variables
    title_font = Font(bold=True, size=20)
    title_alignment = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(
        left=Side(style='thin', color='D9D9D9'), 
        right=Side(style='thin', color='D9D9D9'), 
        top=Side(style='thin', color='D9D9D9'), 
        bottom=Side(style='thin', color='D9D9D9')
    )
    data_font = Font(color="000000")
    data_alignment = Alignment(horizontal='left', vertical='center', indent=1)
    data_border = Border(
        left=Side(style='thin', color='D9D9D9'), 
        right=Side(style='thin', color='D9D9D9'), 
        top=Side(style='thin', color='D9D9D9'), 
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Add a title row
    ws.append(['CMT Data: '])
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    ws.row_dimensions[1].height = excel_title_height

    # Add header row from DataFrame
    headers = combined_df.columns.tolist()
    ws.append(headers)
    for i, cell in enumerate(ws[2]):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
        ws.column_dimensions[chr(65+i)].width = get_column_width(headers[i], min_header_weight)
    ws.row_dimensions[2].height = excel_header_height

    # Add data rows from DataFrame and style them
    for r_idx, row in enumerate(
        dataframe_to_rows(combined_df, index=False, header=False), start=3
    ):
        if all(cell is None or cell == '' for cell in row):
            continue
        ws.append(row)
        ws.row_dimensions[r_idx].height = excel_data_height
        for c_idx, cell in enumerate(row):
            ws_cell = ws.cell(row=r_idx, column=c_idx+1)
            ws_cell.border = data_border
            ws_cell.alignment = data_alignment
            ws_cell.font = data_font
            current_width = ws.column_dimensions[chr(65+c_idx)].width
            cell_width = get_column_width(str(cell), min_header_weight)
            if float(cell_width) > float(current_width) and float(cell_width) < max_header_weight:
                ws.column_dimensions[chr(65+c_idx)].width = cell_width
            elif float(cell_width) > max_header_weight:
                ws.column_dimensions[chr(65+c_idx)].width = max_header_weight

    # Format and save the workbook with a datetime stamp in the filename
    current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    if target_filename and target_filename != "":
        filename = f'{target_filename}_{current_time}.xlsx'
    else:
        filename = f'formatted_excel_{current_time}.xlsx'

    excel_file = BytesIO()
    wb.save(excel_file)

    session_state.excel_data = excel_file.getvalue()
    session_state.download_filename = filename
    session_state.show_download = True